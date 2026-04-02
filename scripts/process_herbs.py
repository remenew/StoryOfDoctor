# -*- coding: utf-8 -*-
"""
从 formula.csv 的"方解（君臣佐使）"列提取药材名
检查 herbs-0401.xlsx 的"name"列是否包含这些药材
如果不包含，则在 name 列最后增加相应药材
"""

import csv
import re
from openpyxl import load_workbook

def extract_herbs_from_formula(formula_text):
    """
    从方解文本中提取药材名
    药材名不包含"君"、"臣"、"佐"、"使"
    """
    if not formula_text:
        return []
    
    # 移除"君"、"臣"、"佐"、"使"等字
    text = re.sub(r'[君臣佐使]', '', formula_text)
    
    # 常见的药材名模式：中文+可能的剂量单位
    # 匹配2-4个汉字的中药材名
    pattern = r'[\u4e00-\u9fa5]{2,4}'
    matches = re.findall(pattern, text)
    
    # 过滤掉常见的非药材词汇
    exclude_words = {'用量', '功效', '主治', '方解', '水煎服', '水二杯', '煎取', '分温', '再服',
                     '一杯', '二杯', '三杯', '温服', '日三服', '日二服', '日一服', '不拘时',
                     '加减', '方歌', '方论', '现代', '运用', '临床', '应用', '医案',
                     '若', '者', '加', '减', '各', '等', '分', '为', '末', '丸', '散', '汤',
                     '方', '药', '症', '证', '病', '治', '疗', '法', '剂', '服', '用'}
    
    herbs = []
    for match in matches:
        if match not in exclude_words and len(match) >= 2:
            herbs.append(match)
    
    return list(set(herbs))  # 去重

def main():
    # 1. 读取 formula.csv
    print("正在读取 formula.csv...")
    formula_herbs = set()
    
    with open('d:/game/med/data/formula.csv', 'r', encoding='gbk') as f:
        reader = csv.DictReader(f)
        for row in reader:
            formula_text = row.get('方解（君臣佐使）', '')
            herbs = extract_herbs_from_formula(formula_text)
            formula_herbs.update(herbs)
    
    print(f"从 formula.csv 提取到 {len(formula_herbs)} 个药材名")
    print(f"药材列表: {sorted(formula_herbs)}")
    
    # 2. 读取 herbs-0401.xlsx
    print("\n正在读取 herbs-0401.xlsx...")
    wb = load_workbook('d:/game/med/data/herbs-0401.xlsx')
    ws = wb.active
    
    # 找到 name 列
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    name_col_idx = None
    for idx, cell_value in enumerate(header_row, 1):
        if cell_value == 'name':
            name_col_idx = idx
            break
    
    if not name_col_idx:
        print("错误：未找到 'name' 列")
        return
    
    print(f"找到 'name' 列在第 {name_col_idx} 列")
    
    # 获取现有的药材名
    existing_herbs = set()
    max_row = ws.max_row
    for row in range(2, max_row + 1):  # 从第2行开始（跳过表头）
        cell_value = ws.cell(row=row, column=name_col_idx).value
        if cell_value:
            existing_herbs.add(str(cell_value).strip())
    
    print(f"herbs-0401.xlsx 中已有 {len(existing_herbs)} 个药材")
    
    # 3. 找出需要添加的药材
    herbs_to_add = formula_herbs - existing_herbs
    
    print(f"\n需要添加 {len(herbs_to_add)} 个新药材:")
    for herb in sorted(herbs_to_add):
        print(f"  - {herb}")
    
    # 4. 添加新药材
    if herbs_to_add:
        current_row = max_row + 1
        for herb in sorted(herbs_to_add):
            ws.cell(row=current_row, column=name_col_idx, value=herb)
            print(f"添加: {herb} -> 第 {current_row} 行")
            current_row += 1
        
        # 保存文件
        wb.save('d:/game/med/data/herbs-0401.xlsx')
        print(f"\n已保存到 herbs-0401.xlsx")
    else:
        print("\n没有需要添加的新药材")
    
    wb.close()
    print("\n处理完成！")

if __name__ == '__main__':
    main()
