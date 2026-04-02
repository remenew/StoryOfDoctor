import openpyxl

wb = openpyxl.load_workbook('D:/game/med/data/herbs-0401.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
desc_col = headers.index('desc') + 1
use_col = headers.index('use') + 1
name_col = headers.index('name') + 1

empty_rows = []
for row in ws.iter_rows(min_row=2):
    desc_val = row[desc_col - 1].value
    if desc_val is None or str(desc_val).strip() == '':
        idx = row[0].row
        name = row[name_col - 1].value
        use_val = row[use_col - 1].value
        empty_rows.append((idx, name, use_val))

with open('D:/game/med/data/empty_herbs.txt', 'w', encoding='utf-8') as f:
    f.write(f"Total empty desc rows: {len(empty_rows)}\n\n")
    for idx, name, use in empty_rows:
        f.write(f"Row {idx}: {name} | use={use}\n")

wb.close()
print("Done! Written to empty_herbs.txt")
