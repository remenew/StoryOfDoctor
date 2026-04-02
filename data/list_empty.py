import openpyxl

wb = openpyxl.load_workbook('D:/game/med/data/herbs-0401.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

# Find column indices (1-based)
desc_col = headers.index('desc') + 1
use_col = headers.index('use') + 1
name_col = headers.index('name') + 1

empty_rows = []
for row in ws.iter_rows(min_row=2):
    desc_val = row[desc_col - 1].value
    if desc_val is None or str(desc_val).strip() == '':
        idx = row[0].row
        name = row[name_col - 1].value
        empty_rows.append((idx, name))

print(f"\nTotal empty desc rows: {len(empty_rows)}")
print("\nHerb names with empty desc:")
for idx, name in empty_rows:
    print(f"  Row {idx}: {name}")

wb.close()
