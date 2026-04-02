import openpyxl

wb = openpyxl.load_workbook('D:/game/med/data/herbs-0401.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
desc_col = headers.index('desc') + 1
use_col = headers.index('use') + 1
name_col = headers.index('name') + 1

# Show a few filled rows to understand format
count = 0
for row in ws.iter_rows(min_row=2):
    desc_val = row[desc_col - 1].value
    if desc_val is not None and str(desc_val).strip() != '':
        name = row[name_col - 1].value
        use_val = row[use_col - 1].value
        with open('D:/game/med/data/format_samples.txt', 'a', encoding='utf-8') as f:
            f.write(f"--- {name} ---\n")
            f.write(f"desc: {desc_val}\n")
            f.write(f"use: {use_val}\n\n")
        count += 1
        if count >= 5:
            break

wb.close()
print(f"Written {count} samples to format_samples.txt")
